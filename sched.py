import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import hashlib

# -------------------------------
# 1️⃣ Load JSON schedule
# -------------------------------
with open("assigned_schedule_updated.json", "r", encoding="utf-8") as f:
    schedule = json.load(f)

# -------------------------------
# 2️⃣ Utility: generate unique color for each clinic
# -------------------------------
def unique_color(name):
    """Generate a pastel-like hex color based on course/clinic name."""
    h = hashlib.md5(name.encode("utf-8")).hexdigest()
    # Take first 6 hex chars and force lighter shades
    r = int(h[0:2], 16) // 2 + 128
    g = int(h[2:4], 16) // 2 + 128
    b = int(h[4:6], 16) // 2 + 128
    return f"{r:02X}{g:02X}{b:02X}"

clinic_colors = {}

# -------------------------------
# 3️⃣ Time utilities
# -------------------------------
def normalize_times(from_str, to_str):
    fmt = "%H:%M"
    from_time = datetime.strptime(from_str.strip(), fmt)
    to_time = datetime.strptime(to_str.strip(), fmt)
    if from_time > to_time:
        from_time, to_time = to_time, from_time
    return from_time.strftime(fmt), to_time.strftime(fmt)

def create_time_grid(start="08:00", end="18:00", step_minutes=30):
    fmt = "%H:%M"
    t_start = datetime.strptime(start, fmt)
    t_end = datetime.strptime(end, fmt)
    times = []
    while t_start < t_end:
        times.append(t_start.strftime(fmt))
        t_start += timedelta(minutes=step_minutes)
    return times

time_grid = create_time_grid()
time_to_col = {t: idx+2 for idx, t in enumerate(time_grid)}  # columns start at B

# -------------------------------
# 4️⃣ Initialize workbook
# -------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Clinics Schedule"

# Column widths
ws.column_dimensions["A"].width = 20
for i in range(len(time_grid)):
    col_letter = chr(66 + i) if i < 25 else f"A{i-25+1}"  # B..Z, then AA, AB...
    ws.column_dimensions[col_letter].width = 15

# Write top time labels
row_idx = 1
for t, col in time_to_col.items():
    ws.cell(row=row_idx, column=col, value=t)
row_idx += 1

# -------------------------------
# 5️⃣ Export schedule by location
# -------------------------------
for day, locations in schedule.items():
    # Day header
    ws.cell(row=row_idx, column=1, value=day)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(time_grid)+1)
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    row_idx += 1

    for loc in ["New Campus", "Old Campus", "CELT"]:
        sessions_in_loc = locations.get(loc, [])
        if not sessions_in_loc:
            continue

        # Location label
        ws.cell(row=row_idx, column=1, value=loc)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1

        # Group by course
        courses = {}
        for s in sessions_in_loc:
            course_name = s["Course"].strip()
            courses.setdefault(course_name, []).append({
                "From": s["From"],
                "To": s["To"],
                "Workers": s["Workers"],
                "Instructor": s["Instructor"]
            })

        # Each course on one row
        for course_name, sessions in courses.items():
            this_row = row_idx  # reserve row for this course

            # Assign unique color if not already assigned
            if course_name not in clinic_colors:
                clinic_colors[course_name] = unique_color(course_name)

            for s in sessions:
                # Find start and end columns
                start_col = min([col for t, col in time_to_col.items() if t >= s["From"]], default=2)
                end_col = max([col for t, col in time_to_col.items() if t < s["To"]], default=start_col)

                # Merge cells for this session in the same row
                ws.merge_cells(start_row=this_row, start_column=start_col, end_row=this_row, end_column=end_col)
                clinic_text = f"{course_name}\n{' / '.join([str(w) for w in s['Workers'] if w])}\n{s['Instructor']}\n{s['From']}-{s["To"]}"
                ws.cell(row=this_row, column=start_col, value=clinic_text)
                ws.cell(row=this_row, column=start_col).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                # Count lines in cell
                num_lines = clinic_text.count("\n") + 1
                ws.row_dimensions[this_row].height = max(ws.row_dimensions[this_row].height or 15, num_lines * 15)  

                # Unique color fill
                fill_color = clinic_colors[course_name]
                for c in range(start_col, end_col+1):
                    ws.cell(row=this_row, column=c).fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid"
                    )

            # after all sessions of that course, move to next row
            row_idx += 1

    # Blank row after each day
    row_idx += 1

# -------------------------------
# 6️⃣ Save Excel
# -------------------------------
wb.save("clinics_schedule_unique_colors.xlsx")
print("✅ Clinics schedule exported to clinics_schedule_unique_colors.xlsx")