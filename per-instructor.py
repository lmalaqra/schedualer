import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime, timedelta
import hashlib
from openpyxl.utils import get_column_letter

# Load JSON files
with open("assigned_schedule_updated.json", "r", encoding="utf-8") as f:
    clinics = json.load(f)

with open("other_schedule.json", "r", encoding="utf-8") as f:
    lectures = json.load(f)

# --- Helper Functions ---
def generate_used_time_slots(entries, interval_minutes=30):
    slots = set()
    for day_entries in entries.values():
        for e in day_entries:
            start, end = e.get("From","").strip(), e.get("To","").strip()
            if start and end:
                start_dt = datetime.strptime(start, "%H:%M")
                end_dt = datetime.strptime(end, "%H:%M")
                cur = start_dt
                while cur < end_dt:
                    slot_end = cur + timedelta(minutes=interval_minutes)
                    slots.add((cur.strftime("%H:%M"), slot_end.strftime("%H:%M")))
                    cur = slot_end
    return sorted(list(slots))

def color_from_string(s, prefix="lec"):
    return hashlib.md5((prefix+s).encode("utf-8")).hexdigest()[:6]

def add_entry(ws, row_info, e, time_slots, label, color_key):
    row = row_info + [""] * len(time_slots)
    ws.append(row)
    row_idx = ws.max_row

    if e.get("From") and e.get("To"):
        start_time = datetime.strptime(e["From"], "%H:%M")
        end_time = datetime.strptime(e["To"], "%H:%M")
        start_col = end_col = None
        for idx, (slot_start, slot_end) in enumerate(time_slots):
            s_dt = datetime.strptime(slot_start, "%H:%M")
            e_dt = datetime.strptime(slot_end, "%H:%M")
            if start_col is None and start_time < e_dt and end_time > s_dt:
                start_col = idx + len(row_info) + 1
            if start_col is not None and start_time < e_dt and end_time > s_dt:
                end_col = idx + len(row_info) + 1
        if start_col is not None:
            if start_col != end_col:
                ws.merge_cells(start_row=row_idx, start_column=start_col,
                               end_row=row_idx, end_column=end_col)
            cell = ws.cell(row=row_idx, column=start_col, value=label)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color=color_key, end_color=color_key, fill_type="solid")

# --- Collect all entries ---
all_entries = []

# Clinics
for day, locs in clinics.items():
    for loc, sessions in locs.items():
        for e in sessions:
            e_copy = e.copy()
            e_copy["Location"] = loc
            all_entries.append((day, e_copy, "clinic"))

# Lectures
for day, lst in lectures.items():
    for e in lst:
        all_entries.append((day, e, "lecture"))

# --- Generate unique time slots ---
lecture_slots = generate_used_time_slots(lectures)
clinic_slots = generate_used_time_slots({
    d: [c for d2, c, t in all_entries if d2 == d and t == "clinic"]
    for d, e, t in all_entries
})
time_slots = sorted(list(set(lecture_slots + clinic_slots)), key=lambda x:(x[0], x[1]))

# --- Group entries per instructor ---
instructors = {}
for day, e, etype in all_entries:
    instructors.setdefault(e["Instructor"], []).append((day, e, etype))

# --- Create Excel Workbook ---
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

for instr, entries in instructors.items():
    ws = wb.create_sheet(title=instr[:30])
    header = ["Day", "Location/Room"] + [f"{s}-{e}" for s,e in time_slots]
    ws.append(header)

    for day, e, etype in entries:
        if etype == "lecture":
            row_info = [day, f"{e.get('Location','')} / {e.get('Room','')}"]
            label = e["Course"]
            color_key = color_from_string(e["Course"], "lec")
        else:
            row_info = [day, f"{e.get('Location','')} / {e.get('Clinic','')}"]
            label = e["Clinic"]
            color_key = color_from_string(e["Clinic"], "cli")
        add_entry(ws, row_info, e, time_slots, label, color_key)

    # --- Formatting ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    for col_idx in range(3, len(time_slots)+3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 6
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   text_rotation=90, wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.alignment is None:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# --- Save Excel ---
wb.save("per_instructor_combined.xlsx")
print("✅ Excel file 'per_instructor_combined.xlsx' generated successfully!")