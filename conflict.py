import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

# -------------------------------
# 1. Send POST request
# -------------------------------
url = "https://zajelbs.najah.edu/servlet/materials"
payload = {
    "b": 10761  # example parameter
    # add other form parameters if needed
}

response = requests.post(url, data=payload)
if response.status_code != 200:
    raise Exception(f"POST request failed with status code {response.status_code}")
response.encoding = "windows-1256"

html_content = response.text


# -------------------------------
# 2. Parse HTML and find lecture table
# -------------------------------
soup = BeautifulSoup(html_content, "html.parser")
tables = soup.find_all("table")

# Expected headers (Arabic) for the lecture table
expected_headers = [
    "المساق/ش", "اسم المساق", "س.م", "الأيام", "الساعة",
    "القاعة", "الحرم", "المتطلبات السابقة", "المدرس", "أرقام مساقات مكافئة"
]

lecture_df = None
for table in tables:
    try:
        df = pd.read_html(str(table), header=0)[0]
        df.columns = [col.strip() for col in df.columns]
        if all(h in df.columns for h in expected_headers):
            lecture_df = df
            break
    except:
        continue

if lecture_df is None:
    raise Exception("Could not find the table with the expected headers")

# -------------------------------
# 3. Build lecture schedule (skip clinics)
# -------------------------------
exclude_keywords = ["عيادة", "عملي", "مختبر"]

def normalize_time(time_str):
    if pd.isna(time_str) or str(time_str).strip() == "":
        return "", ""
    try:
        start, end = time_str.split("-")
        return start.strip(), end.strip()
    except:
        return "", ""

schedule = {}
for _, row in lecture_df.iterrows():
    course_name = str(row["اسم المساق"]).strip() if pd.notna(row["اسم المساق"]) else ""
    
    # Skip clinics
    if any(k in course_name for k in exclude_keywords):
        continue

    day = row["الأيام"] if pd.notna(row["الأيام"]) else "غير محدد"
    start_time, end_time = normalize_time(row["الساعة"] if pd.notna(row["الساعة"]) else "")
    room = str(row["القاعة"]).strip() if pd.notna(row["القاعة"]) else ""
    instructor = str(row["المدرس"]).strip() if pd.notna(row["المدرس"]) else ""

    if not any([course_name, start_time, end_time, room, instructor]):
        continue

    lecture = {
        "Course": course_name,
        "From": start_time,
        "To": end_time,
        "Room": room,
        "Instructor": instructor
    }
    schedule.setdefault(day, []).append(lecture)

# -------------------------------
# 4. Detect conflicts among lectures
# -------------------------------
def parse_time(t):
    return datetime.strptime(t.strip(), "%H:%M")

conflicts = []
for day, lectures in schedule.items():
    rooms = {}
    for lec in lectures:
        rooms.setdefault(lec["Room"], []).append(lec)

    for room, room_lectures in rooms.items():
        room_lectures.sort(key=lambda x: parse_time(x["From"]))
        for i in range(len(room_lectures) - 1):
            lec1 = room_lectures[i]
            lec2 = room_lectures[i + 1]
            start1, end1 = parse_time(lec1["From"]), parse_time(lec1["To"])
            start2, end2 = parse_time(lec2["From"]), parse_time(lec2["To"])
            if start2 < end1:  # Overlap
                conflicts.append([
                    day,
                    room,
                    lec1["Course"], lec1["Instructor"], lec1["From"], lec1["To"],
                    lec2["Course"], lec2["Instructor"], lec2["From"], lec2["To"],
                ])

# -------------------------------
# 5. Export conflicts to Excel
# -------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Conflicts"

headers = [
    "Day", "Room",
    "Course 1", "Instructor 1", "From 1", "To 1",
    "Course 2", "Instructor 2", "From 2", "To 2"
]
ws.append(headers)

for row in conflicts:
    ws.append(row)

# Style headers
for cell in ws[1]:
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

# Auto column width
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max(12, max_length + 2)

wb.save("lecture_conflicts.xlsx")
print("✅ Lecture-only conflict report saved as 'lecture_conflicts.xlsx'")