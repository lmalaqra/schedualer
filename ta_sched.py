import requests
import pandas as pd
from bs4 import BeautifulSoup
import re
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import hashlib

# -------------------------------
# 1️⃣ User inputs
# -------------------------------
b_value = 10761
total_workers = int(input("Enter total number of workers: "))
output_file = "Scheduale.xlsx"

# -------------------------------
# 2️⃣ Fetch schedule from website
# -------------------------------
url = "https://zajelbs.najah.edu/servlet/materials"
payload = {"b": b_value}
response = requests.post(url, data=payload)
if response.status_code != 200:
    raise Exception(f"POST request failed with status code {response.status_code}")

response.encoding = "windows-1256"
html_content = response.text
soup = BeautifulSoup(html_content, "html.parser")
tables = soup.find_all("table")

def normalize_text(s):
    if not s:
        return ""
    s = str(s).replace("\xa0", " ").replace("&nbsp;", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()

# Identify table by headers
key_headers = ["المساق/ش", "اسم المساق", "س.م", "الأيام", "الساعة",
               "القاعة", "الحرم", "المتطلبات السابقة", "المدرس", "أرقام مساقات مكافئة"]
target_table = None
for table in tables:
    first_row = table.find("tr")
    if not first_row:
        continue
    cols = [normalize_text(td.get_text()) for td in first_row.find_all("td")]
    if all(any(kh in c for c in cols) for kh in key_headers):
        target_table = table
        break
if target_table is None:
    raise Exception("Could not find the table with the expected headers")

# Extract rows
rows = []
for tr in target_table.find_all("tr")[1:]:
    cols = tr.find_all("td")
    if len(cols) == 0:
        continue
    row = [normalize_text(td.get_text(separator=" ", strip=True)) for td in cols]
    rows.append(row)

# Create DataFrame
df = pd.DataFrame(rows)
if df.shape[1] >= 12:
    df.columns = ["Image", "Empty1", "المساق/ش", "اسم المساق", "س.م",
                  "الأيام", "الساعة", "القاعة", "الحرم",
                  "المتطلبات السابقة", "المدرس", "أرقام مساقات مكافئة"]
    df = df.drop(columns=["Image", "Empty1"])

# -------------------------------
# 3️⃣ Filter clinics and assign location
# -------------------------------
clinic_keywords = ["عيادة", "عملي", "مختبر"]
old_campus_clinics = [
    "عيادة طب أسنان الأطفال 1",
    "عيادة استعاضة سنية متحركة 4",
    "عيادة جراحة الفم والأسنان والفكين 1",
    "عيادة طب الأسنان التحفظي 5",
    "عيادة مداواة الأسنان اللبية 4",
    "عيادة علم أمراض اللثة 3"
]

def normalize_time(time_str):
    if pd.isna(time_str) or str(time_str).strip() == "":
        return "", ""
    try:
        start, end = time_str.split("-")
        return start.strip(), end.strip()
    except:
        return "", ""

def determine_location(clinic_name):
    if "عملي" in clinic_name or "مختبر" in clinic_name:
        return "New Campus"
    elif clinic_name in old_campus_clinics:
        return "Old Campus"
    else:
        return "CELT"

clinics_schedule = {}
for _, row in df.iterrows():
    row_text = " ".join([str(cell) for cell in row if pd.notna(cell)])
    if not any(k in row_text for k in clinic_keywords):
        continue

    day = row["الأيام"] if pd.notna(row["الأيام"]) else "غير محدد"
    time_str = row["الساعة"] if pd.notna(row["الساعة"]) else ""
    start_time, end_time = normalize_time(time_str)
    room = str(row["القاعة"]).strip() if pd.notna(row["القاعة"]) else ""
    course_name = str(row["اسم المساق"]).strip() if pd.notna(row["اسم المساق"]) else ""
    instructor = str(row["المدرس"]).strip() if pd.notna(row["المدرس"]) else ""

    if not any([course_name, start_time, end_time, room, instructor]):
        continue

    location = determine_location(course_name)
    entry = {"Course": course_name, "From": start_time, "To": end_time,
             "Room": room, "Location": location, "Instructor": instructor}

    if day not in clinics_schedule:
        clinics_schedule[day] = {"New Campus": [], "Old Campus": [], "CELT": []}
    clinics_schedule[day][location].append(entry)

# -------------------------------
# 4️⃣ Worker assignment
# -------------------------------
workers = list(range(1, total_workers+1))
worker_assignments = {w: [] for w in workers}
worker_day_location = {w: {} for w in workers}

def time_to_minutes(t):
    h, m = map(int, t.split(":"))
    return h*60 + m

def is_overlap(start1, end1, start2, end2):
    return not (end1 <= start2 or end2 <= start1)

assigned_schedule = {}
for day, locations in clinics_schedule.items():
    assigned_schedule[day] = {}
    for location, sessions in locations.items():
        sessions_sorted = sorted(sessions, key=lambda s: time_to_minutes(s["From"]))
        assigned_schedule[day][location] = []

        for session in sessions_sorted:
            if location == "New Campus":
                if course_name.strip() == "طب الأسنان التحفظي 1/ عملي":
                    required_workers = 2
                else:
                    required_workers = 1
            else:
    # Old Campus and CELT
                required_workers = 2
            assigned_workers = []

            candidates = [w for w in workers if day in worker_day_location[w] and worker_day_location[w][day] == location]
            candidates += sorted([w for w in workers if w not in candidates], key=lambda w: len(worker_assignments[w]))

            session_start = time_to_minutes(session["From"])
            session_end = time_to_minutes(session["To"])

            for w in candidates:
                if len(assigned_workers) >= required_workers:
                    break
                conflict = any(s_day == day and is_overlap(session_start, session_end, s_start, s_end)
                               for s_day, s_start, s_end in worker_assignments[w])
                if conflict:
                    continue
                worker_assignments[w].append((day, session_start, session_end))
                worker_day_location[w][day] = location
                assigned_workers.append(w)

            while len(assigned_workers) < required_workers:
                assigned_workers.append(None)
                print(f"⚠️ Warning: Not enough workers for {session['Course']} on {day} at {location}")

            session_copy = session.copy()
            session_copy["Workers"] = assigned_workers
            assigned_schedule[day][location].append(session_copy)

# -------------------------------
# 5️⃣ Excel export
# -------------------------------
def unique_color(name):
    h = hashlib.md5(name.encode("utf-8")).hexdigest()
    r = int(h[0:2],16)//2 + 128
    g = int(h[2:4],16)//2 + 128
    b = int(h[4:6],16)//2 + 128
    return f"{r:02X}{g:02X}{b:02X}"

clinic_colors = {}
time_grid = [(datetime.strptime("08:00","%H:%M")+timedelta(minutes=30*i)).strftime("%H:%M") for i in range(20)]
time_to_col = {t: idx+2 for idx,t in enumerate(time_grid)}

wb = Workbook()
ws = wb.active
ws.title = "Clinics Schedule"
ws.column_dimensions["A"].width = 20
for i in range(len(time_grid)):
    ws.column_dimensions[chr(66+i)].width = 15

row_idx = 1
for t,col in time_to_col.items():
    ws.cell(row=row_idx, column=col, value=t)
row_idx +=1

for day, locations in assigned_schedule.items():
    ws.cell(row=row_idx, column=1, value=day)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(time_grid)+1)
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    row_idx +=1

    for loc in ["New Campus","Old Campus","CELT"]:
        sessions_in_loc = locations.get(loc,[])
        if not sessions_in_loc:
            continue

        ws.cell(row=row_idx, column=1, value=loc)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        row_idx +=1

        courses = {}
        for s in sessions_in_loc:
            course_name = s["Course"]
            courses.setdefault(course_name,[]).append(s)

        for course_name,sessions in courses.items():
            this_row = row_idx
            if course_name not in clinic_colors:
                clinic_colors[course_name] = unique_color(course_name)

            for s in sessions:
                start_col = min([col for t,col in time_to_col.items() if t>=s["From"]], default=2)
                end_col = max([col for t,col in time_to_col.items() if t<s["To"]], default=start_col)
                ws.merge_cells(start_row=this_row, start_column=start_col, end_row=this_row, end_column=end_col)
                clinic_text = f"{course_name}\n{' / '.join([str(w) for w in s['Workers'] if w])}\n{s['Instructor']}\n{s['From']}-{s['To']}"
                ws.cell(row=this_row, column=start_col, value=clinic_text)
                ws.cell(row=this_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[this_row].height = max(ws.row_dimensions[this_row].height or 15, clinic_text.count("\n")*15)
                for c in range(start_col,end_col+1):
                    ws.cell(row=this_row, column=c).fill = PatternFill(
                    start_color=clinic_colors[course_name],
                    end_color=clinic_colors[course_name],  # ✅ use course_name
                    fill_type="solid"
)
            row_idx +=1
    row_idx +=1


# -------------------------------
# 7️⃣ Add summary sheet
# -------------------------------
summary_ws = wb.create_sheet(title="Summary")

# Headers
summary_ws.cell(row=1, column=1, value="Worker")
summary_ws.cell(row=1, column=2, value="Total Hours")
summary_ws.cell(row=1, column=3, value="Total Clinics")
summary_ws.cell(row=1, column=4, value="Total Labs/Practicals")

# Calculate totals for each worker
for idx, w in enumerate(workers, start=2):
    total_minutes = 0
    total_clinics = 0
    total_labs = 0

    for day, locs in assigned_schedule.items():
        for loc, sessions in locs.items():
            for s in sessions:
                if w in s["Workers"]:
                    # Add hours
                    start = datetime.strptime(s["From"], "%H:%M")
                    end = datetime.strptime(s["To"], "%H:%M")
                    delta = end - start
                    total_minutes += delta.total_seconds() / 60

                    # Count clinic/lab
                    if "مختبر" in s["Course"] or "عملي" in s["Course"]:
                        total_labs += 1
                    else:
                        total_clinics += 1

    summary_ws.cell(row=idx, column=1, value=w)
    summary_ws.cell(row=idx, column=2, value=round(total_minutes / 60, 2))
    summary_ws.cell(row=idx, column=3, value=total_clinics)
    summary_ws.cell(row=idx, column=4, value=total_labs)

# Optional: adjust column widths
for col in ["A", "B", "C", "D"]:
    summary_ws.column_dimensions[col].width = 20

wb.save(output_file)
print(f"✅ Combined schedule exported to {output_file}")