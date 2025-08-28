import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime, timedelta
import hashlib

# Load JSON
with open("other_schedule.json", "r", encoding="utf-8") as f:
    schedule = json.load(f)

# Function to generate *only used* time slots in 30-min intervals
def generate_used_time_slots(entries, interval_minutes=30):
    used_slots = set()
    for e in entries:
        if e["From"].strip() and e["To"].strip():
            start = datetime.strptime(e["From"], "%H:%M")
            end = datetime.strptime(e["To"], "%H:%M")
            current = start
            while current < end:
                slot_end = current + timedelta(minutes=interval_minutes)
                used_slots.add((current.strftime("%H:%M"), slot_end.strftime("%H:%M")))
                current = slot_end
    return sorted(list(used_slots), key=lambda x: (x[0], x[1]))

# Function to generate a consistent color from a string
def color_from_string(s):
    h = hashlib.md5(s.encode("utf-8")).hexdigest()
    return h[:6]

# Collect all entries
all_entries = [e for day_entries in schedule.values() for e in day_entries]
time_slots = generate_used_time_slots(all_entries)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Other Schedule"

# Header row
header = ["Day", "Location", "Room"] + [f"{start}-{end}" for start, end in time_slots]
ws.append(header)

# Fill rows with merging and coloring
for day, entries in schedule.items():
    grouped = {}
    for e in entries:
        loc = e["Location"]
        room = e["Room"]
        grouped.setdefault(loc, {}).setdefault(room, []).append(e)

    for loc, rooms in grouped.items():
        for room, lectures in rooms.items():
            row = [day, loc, room] + [""] * len(time_slots)
            ws.append(row)
            row_idx = ws.max_row

            for lec in lectures:
                if lec["From"].strip() and lec["To"].strip():
                    start_time = datetime.strptime(lec["From"], "%H:%M")
                    end_time = datetime.strptime(lec["To"], "%H:%M")
                    start_col, end_col = None, None

                    for idx, (slot_start, slot_end) in enumerate(time_slots):
                        slot_start_dt = datetime.strptime(slot_start, "%H:%M")
                        slot_end_dt = datetime.strptime(slot_end, "%H:%M")
                        if start_col is None and start_time < slot_end_dt and end_time > slot_start_dt:
                            start_col = idx + 4
                        if start_col is not None and start_time < slot_end_dt and end_time > slot_start_dt:
                            end_col = idx + 4

                    if start_col is not None and end_col is not None:
                        if start_col != end_col:
                            ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
                        cell = ws.cell(row=row_idx, column=start_col, value=f"{lec['Course']} ({lec['Instructor']})")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        fill_color = color_from_string(lec["Course"])
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# Adjust column widths
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 12
for col_idx in range(4, len(time_slots) + 4):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 6  # narrow since headers are rotated

# Rotate header time labels
for cell in ws[1]:  # first row (header row)
    cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)

# Center align everything else
for row in ws.iter_rows(min_row=2):
    for cell in row:
        if cell.alignment is None:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Save Excel
wb.save("other_schedule_time_compact.xlsx")
print("✅ Excel file 'other_schedule_time_compact.xlsx' generated successfully.")