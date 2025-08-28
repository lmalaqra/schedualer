import pandas as pd
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import hashlib

# ===============================
# Load single input Excel
# ===============================
import requests
import pandas as pd
from bs4 import BeautifulSoup
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import hashlib

# ===============================
# 1. Fetch the table from website
# ===============================
url = "https://zajelbs.najah.edu/servlet/materials"
payload = {
    "b": 10761  # adjust as needed
}

response = requests.post(url, data=payload)
if response.status_code != 200:
    raise Exception(f"POST request failed with status code {response.status_code}")

# Use correct Arabic encoding
response.encoding = "windows-1256"
html_content = response.text

soup = BeautifulSoup(html_content, "html.parser")
tables = soup.find_all("table")

# Normalize text
def normalize_text(s):
    if not s:
        return ""
    s = str(s)
    s = s.replace("\xa0", " ").replace("&nbsp;", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()

# Identify table by headers
key_headers = [
    "المساق/ش", "اسم المساق", "س.م", "الأيام", "الساعة",
    "القاعة", "الحرم", "المتطلبات السابقة", "المدرس", "أرقام مساقات مكافئة"
]

target_table = None
for table in tables:
    first_row = table.find("tr")
    if not first_row:
        continue
    cols = [normalize_text(td.get_text()) for td in first_row.find_all("td")]
    if all(any(kh in c for c in cols) for kh in key_headers):
        target_table = table
        break

if target_table is None:
    raise Exception("Could not find the table with the expected headers")

# Extract rows
rows = []
for tr in target_table.find_all("tr")[1:]:
    cols = tr.find_all("td")
    if len(cols) == 0:
        continue
    row = [normalize_text(td.get_text(separator=" ", strip=True)) for td in cols]
    rows.append(row)

# Create DataFrame
df = pd.DataFrame(rows)
if df.shape[1] >= 12:
    df.columns = [
        "Image", "Empty1", "المساق/ش", "اسم المساق", "س.م",
        "الأيام", "الساعة", "القاعة", "الحرم",
        "المتطلبات السابقة", "المدرس", "أرقام مساقات مكافئة"
    ]
    df = df.drop(columns=["Image", "Empty1"])  # remove unused columns

# ===============================
# 2. Split Clinics / Lectures
# ===============================
exclude_keywords = ["عيادة", "عملي", "مختبر"]

clinics_rows = []
lectures_rows = []

for _, row in df.iterrows():
    row_text = " ".join([str(cell) for cell in row if pd.notna(cell)])
    if any(k in row_text for k in exclude_keywords):
        clinics_rows.append(row)
    else:
        lectures_rows.append(row)

clinics_df = pd.DataFrame(clinics_rows, columns=df.columns)
lectures_df = pd.DataFrame(lectures_rows, columns=df.columns)

# ===============================
# 3. Process Clinics
# ===============================
assigned_schedule = {}

def normalize_time(time_str):
    if pd.isna(time_str) or str(time_str).strip() == "":
        return "", ""
    try:
        start, end = time_str.split("-")
        return start.strip(), end.strip()
    except:
        return "", ""

for _, row in clinics_df.iterrows():
    day = row["الأيام"] if pd.notna(row["الأيام"]) else "غير محدد"
    time_str = row["الساعة"] if pd.notna(row["الساعة"]) else ""
    start_time, end_time = normalize_time(time_str)
    room = str(row["القاعة"]).strip() if pd.notna(row["القاعة"]) else ""
    location = str(row["الحرم"]).strip() if pd.notna(row["الحرم"]) else ""
    course_name = str(row["اسم المساق"]).strip() if pd.notna(row["اسم المساق"]) else ""
    instructor = str(row["المدرس"]).strip() if pd.notna(row["المدرس"]) else ""

    if not any([course_name, start_time, end_time, room, location, instructor]):
        continue

    entry = {
        "Course": course_name,
        "From": start_time,
        "To": end_time,
        "Room": room,
        "Location": location,
        "Instructor": instructor
    }

    assigned_schedule.setdefault(day, []).append(entry)

# ===============================
# 4. Process Lectures
# ===============================
other_schedule = {}

for _, row in lectures_df.iterrows():
    day = row["الأيام"] if pd.notna(row["الأيام"]) else "غير محدد"
    time_str = row["الساعة"] if pd.notna(row["الساعة"]) else ""
    start_time, end_time = normalize_time(time_str)
    room = str(row["القاعة"]).strip() if pd.notna(row["القاعة"]) else ""
    location = str(row["الحرم"]).strip() if pd.notna(row["الحرم"]) else ""
    course_name = str(row["اسم المساق"]).strip() if pd.notna(row["اسم المساق"]) else ""
    instructor = str(row["المدرس"]).strip() if pd.notna(row["المدرس"]) else ""

    if not any([course_name, start_time, end_time, room, location, instructor]):
        continue

    entry = {
        "Course": course_name,
        "From": start_time,
        "To": end_time,
        "Room": room,
        "Location": location,
        "Instructor": instructor
    }

    other_schedule.setdefault(day, []).append(entry)

# ===============================
# 5. Export to Excel (same formatting as before)
# ===============================
wb = Workbook()
# Clinics and Lectures export logic can be copied from your existing code
# (ws1 = wb.active -> Clinics, ws2 = wb.create_sheet("Lectures"))

print("✅ Data fetched from website and processed. Ready for Excel export.")
# Keywords to detect clinics
exclude_keywords = ["عيادة", "عملي", "مختبر"]

clinics_rows = []
lectures_rows = []

for _, row in df.iterrows():
    row_text = " ".join([str(cell) for cell in row if pd.notna(cell)])
    if any(k in row_text for k in exclude_keywords):
        clinics_rows.append(row)
    else:
        lectures_rows.append(row)

clinics_df = pd.DataFrame(clinics_rows, columns=df.columns)
lectures_df = pd.DataFrame(lectures_rows, columns=df.columns)

# ===============================
# Process Clinics Schedule
# ===============================
assigned_schedule = {}

def normalize_time(time_str):
    if pd.isna(time_str) or str(time_str).strip() == "":
        return "", ""
    try:
        start, end = time_str.split("-")
        return start.strip(), end.strip()
    except:
        return "", ""

for _, row in clinics_df.iterrows():
    day = row[3] if pd.notna(row[3]) else "غير محدد"
    time_str = row[4] if pd.notna(row[4]) else ""
    start_time, end_time = normalize_time(time_str)
    room = str(row[5]).strip() if pd.notna(row[5]) else ""
    location = str(row[6]).strip() if pd.notna(row[6]) else ""
    course_name = str(row[1]).strip() if pd.notna(row[1]) else ""
    instructor = str(row[8]).strip() if pd.notna(row[8]) else ""

    if not any([course_name, start_time, end_time, room, location, instructor]):
        continue

    entry = {
        "Course": course_name,
        "From": start_time,
        "To": end_time,
        "Room": room,
        "Location": location,
        "Instructor": instructor
    }

    if day not in assigned_schedule:
        assigned_schedule[day] = []
    assigned_schedule[day].append(entry)

# ===============================
# Process Lectures Schedule
# ===============================
other_schedule = {}

for _, row in lectures_df.iterrows():
    day = row[3] if pd.notna(row[3]) else "غير محدد"
    time_str = row[4] if pd.notna(row[4]) else ""
    start_time, end_time = normalize_time(time_str)
    room = str(row[5]).strip() if pd.notna(row[5]) else ""
    location = str(row[6]).strip() if pd.notna(row[6]) else ""
    course_name = str(row[1]).strip() if pd.notna(row[1]) else ""
    instructor = str(row[8]).strip() if pd.notna(row[8]) else ""

    if not any([course_name, start_time, end_time, room, location, instructor]):
        continue

    entry = {
        "Course": course_name,
        "From": start_time,
        "To": end_time,
        "Room": room,
        "Location": location,
        "Instructor": instructor
    }

    if day not in other_schedule:
        other_schedule[day] = []
    other_schedule[day].append(entry)

# ===============================
# Export to Excel
# ===============================
wb = Workbook()

# -------- Clinics Sheet (Blocked Format) --------
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime, timedelta

ws1 = wb.active
ws1.title = "Clinics"

# Colors
colors = {
    "الجديد": "90EE90",
    "القديم": "ADD8E6",
    "CELT": "FFD700",
    "Lab/Practical": "FFB6C1"
}

# Create time grid
def create_time_grid(start="08:00", end="18:00", step_minutes=30):
    fmt = "%H:%M"
    t_start = datetime.strptime(start, fmt)
    t_end = datetime.strptime(end, fmt)
    times = []
    while t_start < t_end:
        times.append(t_start.strftime(fmt))
        t_start += timedelta(minutes=step_minutes)
    return times

time_grid = create_time_grid()
time_to_col = {t: idx+2 for idx, t in enumerate(time_grid)}  # columns start at B

# Column widths
ws1.column_dimensions["A"].width = 25
for i in range(len(time_grid)):
    col_letter = get_column_letter(i + 2)
    ws1.column_dimensions[col_letter].width = 15

row_idx = 1

# Optional: write top time labels
for t, col in time_to_col.items():
    ws1.cell(row=row_idx, column=col, value=t)
row_idx += 1

# Group entries by day and location
for day, entries in assigned_schedule.items():
    ws1.cell(row=row_idx, column=1, value=day)
    ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(time_grid)+1)
    ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    row_idx += 1

    # Group by location (الجديد, القديم, CELT)
    locations = sorted(set(e["Location"] for e in entries if e["Location"]))
    for loc in locations:
        sessions_in_loc = [e for e in entries if e["Location"] == loc]
        # Group by course
        clinics_grouped = {}
        for s in sessions_in_loc:
            clinics_grouped.setdefault(s["Course"], []).append(s)

        for clinic_name, sessions in clinics_grouped.items():
            ws1.cell(row=row_idx, column=1, value=clinic_name)

            for s in sessions:
                if not s["From"] or not s["To"]:
                    continue
                start_col = min([col for t, col in time_to_col.items() if t >= s["From"]], default=2)
                end_col = max([col for t, col in time_to_col.items() if t < s["To"]], default=start_col)

                # Merge cells for session
                ws1.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
                ws1.cell(row=row_idx, column=start_col, value=f"{s['From']} - {s['To']}")

                ws1.merge_cells(start_row=row_idx+1, start_column=start_col, end_row=row_idx+1, end_column=end_col)
                ws1.cell(row=row_idx+1, column=start_col, value=s.get("Instructor",""))

                ws1.merge_cells(start_row=row_idx+2, start_column=start_col, end_row=row_idx+2, end_column=end_col)
                ws1.cell(row=row_idx+2, column=start_col, value="")  # Workers column can be skipped or added

                # Fill color
                fill_color = colors.get(loc, "FFFFFF")
                if "عملي" in clinic_name or "مختبر" in clinic_name:
                    fill_color = colors["Lab/Practical"]

                for r in range(row_idx, row_idx+3):
                    for c in range(start_col, end_col+1):
                        ws1.cell(r, column=c).fill = PatternFill(start_color=fill_color,
                                                                 end_color=fill_color,
                                                                 fill_type="solid")
                        ws1.cell(r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 3
    row_idx += 1
# -------- Lectures Sheet --------
ws2 = wb.create_sheet("Lectures")

# Helper: generate only used 30-min slots
def generate_used_time_slots(entries, interval_minutes=30):
    used_slots = set()
    for e in entries:
        if e["From"].strip() and e["To"].strip():
            start = datetime.strptime(e["From"], "%H:%M")
            end = datetime.strptime(e["To"], "%H:%M")
            current = start
            while current < end:
                slot_end = current + timedelta(minutes=interval_minutes)
                used_slots.add((current.strftime("%H:%M"), slot_end.strftime("%H:%M")))
                current = slot_end
    return sorted(list(used_slots), key=lambda x: (x[0], x[1]))

def color_from_string(s):
    h = hashlib.md5(s.encode("utf-8")).hexdigest()
    return h[:6]

all_entries = [e for day_entries in other_schedule.values() for e in day_entries]
time_slots = generate_used_time_slots(all_entries)

header = ["Day", "Location", "Room"] + [f"{start}-{end}" for start, end in time_slots]
ws2.append(header)

for day, entries in other_schedule.items():
    grouped = {}
    for e in entries:
        loc = e["Location"]
        room = e["Room"]
        grouped.setdefault(loc, {}).setdefault(room, []).append(e)

    for loc, rooms in grouped.items():
        for room, lectures in rooms.items():
            row = [day, loc, room] + [""] * len(time_slots)
            ws2.append(row)
            row_idx = ws2.max_row

            for lec in lectures:
                if lec["From"].strip() and lec["To"].strip():
                    start_time = datetime.strptime(lec["From"], "%H:%M")
                    end_time = datetime.strptime(lec["To"], "%H:%M")
                    start_col, end_col = None, None

                    for idx, (slot_start, slot_end) in enumerate(time_slots):
                        slot_start_dt = datetime.strptime(slot_start, "%H:%M")
                        slot_end_dt = datetime.strptime(slot_end, "%H:%M")
                        if start_col is None and start_time < slot_end_dt and end_time > slot_start_dt:
                            start_col = idx + 4
                        if start_col is not None and start_time < slot_end_dt and end_time > slot_start_dt:
                            end_col = idx + 4

                    if start_col is not None and end_col is not None:
                        if start_col != end_col:
                            ws2.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
                        cell = ws2.cell(row=row_idx, column=start_col, value=f"{lec['Course']} ({lec['Instructor']})")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        fill_color = color_from_string(lec["Course"])
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# Adjust column widths
ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 12
for col_idx in range(4, len(time_slots) + 4):
    col_letter = get_column_letter(col_idx)
    ws2.column_dimensions[col_letter].width = 6

for cell in ws2[1]:
    cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)

for row in ws2.iter_rows(min_row=2):
    for cell in row:
        if cell.alignment is None:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Save final Excel
wb.save("master_schedule.xlsx")
print("✅ Combined schedule saved to master_schedule.xlsx")